import openpyxl
import json

load_book = openpyxl.load_workbook('easy_japanese.xlsx')
sheet = load_book['Sheet1']
json_path_ja_en = '/home/yukaalive/2025workspace/task_vectors/21_icl_task_vectors/data/translation/ja_en_multi.json'
json_path_en_ja = '/home/yukaalive/2025workspace/task_vectors/21_icl_task_vectors/data/translation/ja_en_multi.json'

translation_dict_ja_en = {}
translation_dict_en_ja = {}

for i in range(2, 502):
    japanese = sheet.cell(row=i, column=3).value
    english = sheet.cell(row=i, column=4).value
    
    # 日本語をキー、英語を値として辞書に追加
    if japanese and english:  # None値をスキップ
        translation_dict_ja_en[japanese] = english

for i in range(2, 502):
    japanese = sheet.cell(row=i, column=3).value
    english = sheet.cell(row=i, column=4).value
    
    # 英語をキー、日本語を値として辞書に追加
    if japanese and english:  # None値をスキップ
        translation_dict_en_ja[english] = japanese

with open(json_path_ja_en, mode='w', encoding='utf-8') as f1:
    f1.write(json.dumps(translation_dict_ja_en, ensure_ascii=False, indent=4))

with open(json_path_en_ja, mode='w', encoding='utf-8') as f2:
    f2.write(json.dumps(translation_dict_en_ja, ensure_ascii=False, indent=4))